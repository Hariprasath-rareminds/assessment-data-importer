#!/usr/bin/env python3
"""Convert Google Form assessment responses into PostgreSQL INSERT SQL.

Inputs:
  - Assessment Answers.xlsx
  - question_bank.xlsx

Output:
  - output/assessment_import.sql
  - output/validation_report.csv

Target tables ONLY:
  - personal_assessment_attempts
  - adaptive_aptitude_sessions
  - adaptive_aptitude_responses
  - adaptive_aptitude_results

Important rules:
  * Existing learner IDs are resolved by email in SQL; learners are never created.
  * Existing question UUIDs are reused; question UUIDs are never generated.
  * New UUIDs are generated only for attempt/session/response/result rows.
  * General Aptitude is stored in adaptive tables, matching the sample data model.
  * Non-adaptive stages are stored in personal_assessment_attempts.all_responses.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


FORM_SHEETS = {
    "Big5": "bigfive",
    "Riasec": "riasec",
    "Employability Assessment ": "employability",
    "Work Values": "values",
    "Genearl Apt": "adaptive_aptitude",
    "MBA Domain knowledge": "mba_knowledge",
    "mba apt": "mba_aptitude",
}

IDENTITY_HEADERS = {
    "timestamp",
    "email address",
    "score",
    "full name",
    "name",
    "email id",
    "contact",
    "contact ",
    "stream",
}

LIKERT_MAPS = {
    "bigfive": {
        "very inaccurate": 1,
        "moderately inaccurate": 2,
        "neither": 3,
        "moderately accurate": 4,
        "very accurate": 5,
    },
    "riasec": {
        "strongly dislike": 1,
        "dislike": 2,
        "neutral": 3,
        "like": 4,
        "strongly like": 5,
    },
    "values": {
        "not important": 1,
        "slightly important": 2,
        "moderately important": 3,
        "very important": 4,
        "extremely important": 5,
        "extremely impotant": 5,  # typo in Google Form responses
    },
    "employability": {
        "not like me": 1,
        "slightly like me": 2,
        "somewhat like me": 3,
        "mostly like me": 4,
        "very much like me": 5,
    },
}


# Verified mappings from the supplied assessment/question-bank review.
# These are used only when textual normalization is insufficient.
VERIFIED_STATIC_UUIDS = {
    "work should fit my life priorities.": "faac271d-0ade-4fee-ae0d-976793779f88",
    "work should fit my life priorities. 2": "faac271d-0ade-4fee-ae0d-976793779f88",
    "evaluate: (60 + 89) × 6 − 3": "1b483fbb-e262-4c32-af03-8d67693573e6",
    "evaluate: (50 + 48) × 6 − 2": "1bf2d8cf-14ab-4334-bffc-84392410d8f1",
    "data table: q1=184, q2=145, q3=123, q4=163. what is the total for the year?": "0fb8bf96-4419-47f3-8714-676fc9e12682",
    "data table: q1=124, q2=157, q3=113, q4=80. what is the total for the year?": "3020b398-f051-4071-a4db-b5282afb87f8",
    "if log₁₀(1,000,000) = k, what is k?": "196980e4-5041-437a-82c0-85414bdec8eb",
    "in a 2×2 grid, bottom-right equals the sum of the other three. if the grid is [[21,9],[21,?]], find ?.": "0da21e72-9077-47e6-9dd6-988bc1e48adf",
    "in a 2×2 grid, bottom-right equals the sum of the other three. if the grid is [[30,19],[9,?]], find ?.": "1ce56a05-48cf-4116-abfc-d72d32d72577",
    "in a 2×2 grid, bottom-right equals the sum of the other three. if the grid is [[32,10],[20,?]], find ?.": "2354c56b-4c26-4ec2-93e8-76c055d57a3f",
    "in a group, |a|=44, |b|=42, and |a∩b|=18. what is |a∪b|?": "005b3803-b53f-43a6-bf42-9502540eb9ae",
    "in a group, |a|=42, |b|=38, and |a∩b|=26. what is |a∪b|?": "286c5cad-a097-455d-b168-e9c7449c4ec2",
    "category values: x=123, y=93, z=147. what is x's share x/(x+y+z)? round to 4 decimals.": "0e8e698d-ede3-4644-a889-83d778febdee",
    "index calculation: base value=144, current value=121. compute index=(current/base)×100. round to 2 decimals.": "149ce6cc-9b31-45e4-968a-7c728edd1779",
    "quarter values: q1=152, q4=129. compute proportional change (q4−q1)/q1. round to 4 decimals.": "11a3f4bd-2181-45ae-8006-070b7a89d5c3",
    "category values: x=137, y=90, z=128. what is x's share x/(x+y+z)? round to 4 decimals.": "1cc561a1-33ea-4a9a-8b20-52b57db34825",
    "a dataset has mean 37 and standard deviation 11. compute coefficient of variation (sd/mean). round to 4 decimals.": "1cc9ca64-657a-4fb5-a6b2-185116412d42",
    "how many ways can 2 items be chosen from 12 distinct items?": "3738d88c-3e30-483c-8c32-d2ed7f384d01",
    "if log₅(125) = k, what is k?": "3a2a6833-8c83-493d-af58-76d08222ef64",
    "data table: q1=115, q2=158, q3=90, q4=137. what is the total for the year?": "4238c753-3041-4d96-a05c-69134fa6a173",
    "quarter values: q1=110, q4=145. compute proportional change (q4−q1)/q1. round to 4 decimals.": "3e179b76-3625-42fc-94fb-2aaad9468e2d",
}

# Existing sample uses these phase names. Google Forms did not actually run the
# adaptive engine, so phase assignment is only for schema/shape compatibility.
def phase_for_sequence(sequence: int) -> str:
    if sequence <= 8:
        return "diagnostic_screener"
    if sequence <= 44:
        return "adaptive_core"
    return "stability_confirmation"


@dataclass
class StaticQuestion:
    id: str
    text: str
    question_type: str
    options: Dict[str, Any]
    correct_answer: Optional[str]
    metadata: Dict[str, Any]
    description: Optional[str]


@dataclass
class AiQuestion:
    id: str
    text: str
    options: List[Any]
    correct_answer: Optional[str]
    category: Optional[str]
    skill_tag: Optional[str]
    difficulty: Optional[str]
    question_type: str


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def repair_mojibake(text: str) -> str:
    replacements = {
        "Ã—": "×",
        "âˆ’": "−",
        "â€“": "-",
        "â€”": "-",
        "â€™": "'",
        "â€œ": '"',
        "â€": '"',
        "Â₹": "₹",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)
    return text


def normalize_question(value: Any) -> str:
    """Strict-but-tolerant normalization for VERIFIED question matching."""
    s = repair_mojibake(as_text(value))
    s = unicodedata.normalize("NFKC", s).lower().strip()
    s = s.replace("’", "'").replace("‘", "'")
    s = s.replace("“", '"').replace("”", '"')
    s = s.replace("×", "*").replace("−", "-").replace("–", "-").replace("—", "-")
    s = s.replace("₁₀", "10").replace("₅", "5")
    s = s.replace("₹", "").replace("$", "")
    s = s.replace(",", "")
    # SQL-dump escaping should not affect semantic text.
    s = s.replace("''", "'")
    # Remove Google Forms duplicate-column suffix.
    s = re.sub(r"\s+\.?(?:2)$", "", s)
    # Remove passage/form wrappers but keep the actual content.
    s = re.sub(r"\bread the passage and answer the question\s*:?", "", s)
    s = re.sub(r"\bpassage\s*:\s*", "", s)
    s = re.sub(r"\bquestion\s*:\s*", "", s)
    # Ignore punctuation/whitespace differences.
    s = re.sub(r"[^a-z0-9%+*/=<>?\-\[\]().| ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_answer(value: Any) -> str:
    s = repair_mojibake(as_text(value))
    s = unicodedata.normalize("NFKC", s).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_answer_cmp(value: Any) -> str:
    s = normalize_answer(value).lower()
    s = s.replace("₹", "").replace("$", "").replace(",", "")
    s = s.replace("−", "-").replace("–", "-").replace("—", "-")
    s = re.sub(r"^[a-d][.)]\s*", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Canonicalize numeric values so Excel 4.0 matches option text "4".
    try:
        n = float(s)
        return format(n, ".12g")
    except ValueError:
        return s


def json_load(value: Any, fallback: Any) -> Any:
    if value in (None, ""):
        return fallback
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(str(value))
    except Exception:
        return fallback


def sql_smallint_array(values: Iterable[Any]) -> str:
    """Serialize a Python sequence as PostgreSQL smallint[]."""
    ints = [str(int(v)) for v in values]
    return "ARRAY[" + ",".join(ints) + "]::smallint[]"


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, (dict, list)):
        raw = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
        return "'" + raw.replace("'", "''") + "'::jsonb"
    if isinstance(value, datetime):
        return "'" + value.isoformat() + "'::timestamptz"
    return "'" + str(value).replace("'", "''") + "'"


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def workbook_rows(ws) -> Iterable[Dict[str, Any]]:
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return
    keys = [as_text(x) for x in headers]
    for row in rows:
        if not row or not any(v not in (None, "") for v in row):
            continue
        yield {keys[i]: row[i] if i < len(row) else None for i in range(len(keys))}


def load_reference_bank(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)

    static_by_norm: Dict[str, List[StaticQuestion]] = defaultdict(list)
    static_by_id: Dict[str, StaticQuestion] = {}
    ws = wb["personal_assessment_questions"]
    for row in workbook_rows(ws):
        qid = as_text(row.get("id"))
        text = as_text(row.get("question_text"))
        if not qid or not text:
            continue
        q = StaticQuestion(
            id=qid,
            text=text,
            question_type=as_text(row.get("question_type")),
            options=json_load(row.get("options"), {}),
            correct_answer=as_text(json_load(row.get("correct_answer"), None)) or None,
            metadata=json_load(row.get("metadata"), {}),
            description=as_text(row.get("description")) or None,
        )
        static_by_norm[normalize_question(text)].append(q)
        static_by_id[qid] = q

    ai_by_norm: Dict[str, List[AiQuestion]] = defaultdict(list)
    for row in workbook_rows(wb["career_assessment_ai_questions"]):
        parent_type = as_text(row.get("question_type"))
        for item in json_load(row.get("questions"), []):
            if not isinstance(item, dict):
                continue
            qid = as_text(item.get("uuid") or item.get("id"))
            text = as_text(item.get("question"))
            if not qid or not text:
                continue
            q = AiQuestion(
                id=qid,
                text=text,
                options=item.get("options") or [],
                correct_answer=as_text(item.get("correct_answer")) or None,
                category=as_text(item.get("category")) or None,
                skill_tag=as_text(item.get("skill_tag")) or None,
                difficulty=as_text(item.get("difficulty")) or None,
                question_type=parent_type,
            )
            ai_by_norm[normalize_question(text)].append(q)

    return static_by_norm, static_by_id, ai_by_norm


def load_ai_questions_from_sql(path: Optional[Path], ai_by_norm: Dict[str, List[AiQuestion]]) -> None:
    if not path or not path.exists():
        return
    text = path.read_text(encoding="utf-8", errors="replace")
    # Supabase dump stores the questions JSON as a single quoted SQL string.
    for m in re.finditer(r", '(\[\{.*?\}\])', '\d{4}-", text, flags=re.S):
        raw = m.group(1).replace("''", "'")
        try:
            items = json.loads(raw)
        except json.JSONDecodeError:
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            qid = as_text(item.get("uuid") or item.get("id"))
            qtext = as_text(item.get("question"))
            if not qid or not qtext:
                continue
            q = AiQuestion(
                id=qid, text=qtext, options=item.get("options") or [],
                correct_answer=as_text(item.get("correct_answer")) or None,
                category=as_text(item.get("category")) or None,
                skill_tag=as_text(item.get("skill_tag")) or None,
                difficulty=as_text(item.get("difficulty")) or None,
                question_type="aptitude",
            )
            key = normalize_question(qtext)
            if not any(existing.id == q.id for existing in ai_by_norm[key]):
                ai_by_norm[key].append(q)


def build_alias_index(reference_norms: Iterable[str]) -> Dict[str, str]:
    """Manual aliases for known wrapper differences verified in this dataset."""
    aliases: Dict[str, str] = {}
    for key in reference_norms:
        # Some DB questions append parenthetical formula hints omitted by Forms.
        base = re.sub(r"\s*\([^)]*(?:round|q1\+q2|current/base|sd/mean)[^)]*\)\s*", " ", key)
        base = re.sub(r"\s+", " ", base).strip()
        aliases.setdefault(base, key)
    return aliases


def resolve_question(header: str, lookup: Dict[str, List[Any]], aliases: Dict[str, str], by_id: Optional[Dict[str, Any]] = None) -> Optional[Any]:
    raw_key = as_text(header).strip().lower()
    if by_id:
        verified_id = VERIFIED_STATIC_UUIDS.get(raw_key)
        if verified_id and verified_id in by_id:
            return by_id[verified_id]
    key = normalize_question(header)
    candidates = lookup.get(key)
    if candidates and len(candidates) == 1:
        return candidates[0]
    aliased = aliases.get(key)
    if aliased:
        candidates = lookup.get(aliased)
        if candidates and len(candidates) == 1:
            return candidates[0]
    # Conservative fallback: containment only when one side is a long extension
    # of the other and there is exactly one candidate.
    matches = []
    for ref_key, qs in lookup.items():
        if len(key) >= 24 and len(ref_key) >= 24 and (key in ref_key or ref_key in key):
            matches.extend(qs)
    uniq = {getattr(q, "id", id(q)): q for q in matches}
    return next(iter(uniq.values())) if len(uniq) == 1 else None


def strip_sjt_suffix(header: str) -> Tuple[str, Optional[str]]:
    m = re.search(r"\s*\[(best|worst)\]\s*(?:\d+)?\s*$", header, flags=re.I)
    if not m:
        return header, None
    kind = m.group(1).lower()
    return header[: m.start()].strip(), kind


def selected_option_key(options: Dict[str, Any], answer: Any) -> Optional[str]:
    cmp = norm_answer_cmp(answer)
    for key, value in options.items():
        if norm_answer_cmp(value) == cmp:
            return str(key).strip('"')
    # Google Forms sometimes stores an option letter directly.
    direct = normalize_answer(answer).upper()
    if direct in options:
        return direct
    return None


def _answer_candidates_for_match(value: Any) -> List[str]:
    """Return normalized comparison candidates without changing stored output."""
    raw = normalize_answer(value)
    candidates = {norm_answer_cmp(raw)}

    # Excel may expose percentage options as decimals (0.2) while the
    # question bank stores the displayed option (20%). Compare both forms.
    compact = raw.replace(",", "").strip()
    try:
        if compact.endswith("%"):
            numeric = float(compact[:-1]) / 100
            candidates.add(format(numeric, ".12g"))
        else:
            numeric = float(compact)
            if 0 <= numeric <= 1:
                percent = numeric * 100
                candidates.add(format(percent, ".12g") + "%")
    except ValueError:
        pass

    return list(candidates)


def ai_answer_for_storage(answer: Any, options: List[Any]) -> str:
    """Map a Google/Excel answer to the exact original question-bank option."""
    raw = normalize_answer(answer)
    answer_candidates = set(_answer_candidates_for_match(raw))

    for opt in options:
        opt_text = normalize_answer(opt)
        option_candidates = set(_answer_candidates_for_match(opt_text))
        if answer_candidates & option_candidates:
            # Store the exact bank option, including A./B./C./D. and %.
            return opt_text

    # If no option matches, preserve the source answer for validation/debugging.
    return raw


def infer_subtag(q: StaticQuestion) -> str:
    md = q.metadata or {}
    for key in ("subtag", "skill_tag", "dimension_name", "category"):
        if md.get(key):
            return str(md[key]).strip().lower().replace(" ", "_")
    dim = str(md.get("dimension", "")).upper()
    mapping = {
        "NR": "numerical_reasoning",
        "LR": "logical_reasoning",
        "DI": "data_interpretation",
        "PR": "pattern_recognition",
        "VR": "verbal_reasoning",
        "SR": "spatial_reasoning",
        "PS": "numerical_reasoning",
        "ST": "logical_reasoning",
    }
    if dim in mapping:
        return mapping[dim]
    desc = (q.description or "").lower()
    if "data interpretation" in desc:
        return "data_interpretation"
    if "pattern" in desc:
        return "pattern_recognition"
    if "logical" in desc or "systems thinking" in desc:
        return "logical_reasoning"
    return "numerical_reasoning"


def infer_difficulty(q: StaticQuestion) -> int:
    md = q.metadata or {}
    for key in ("difficulty_rank", "difficulty", "level"):
        try:
            v = int(md.get(key))
            return max(1, min(5, v))
        except Exception:
            pass
    return 1


def score_summary(responses: List[Dict[str, Any]]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    by_diff: Dict[str, Dict[str, float]] = {}
    by_subtag: Dict[str, Dict[str, float]] = {}
    for item in responses:
        d = str(item["difficulty_at_time"])
        s = item["subtag"]
        for bucket, key in ((by_diff, d), (by_subtag, s)):
            b = bucket.setdefault(key, {"total": 0, "correct": 0, "accuracy": 0})
            b["total"] += 1
            b["correct"] += 1 if item["is_correct"] else 0
    for bucket in (by_diff, by_subtag):
        for b in bucket.values():
            b["accuracy"] = (b["correct"] / b["total"] * 100) if b["total"] else 0
    # Match sample's difficulty keys 1..5 even if unused.
    for d in range(1, 6):
        by_diff.setdefault(str(d), {"total": 0, "correct": 0, "accuracy": 0})
    return by_diff, by_subtag


def confidence_from_accuracy(acc: float) -> str:
    if acc >= 75:
        return "high"
    if acc >= 45:
        return "medium"
    return "low"


def aptitude_level_from_accuracy(acc: float) -> int:
    # Migration fallback only. Replace with production scoring function if the
    # application's exact adaptive scoring algorithm is available.
    if acc >= 85:
        return 5
    if acc >= 70:
        return 4
    if acc >= 55:
        return 3
    if acc >= 35:
        return 2
    return 1


def tier_from_level(level: int) -> str:
    return {1: "M", 2: "M", 3: "H", 4: "H", 5: "H"}.get(level, "M")


def find_email(row: Dict[str, Any]) -> str:
    for key in ("Email Address", "Email ID", "Email Id"):
        value = as_text(row.get(key)).lower()
        if value and "@" in value:
            return value
    return ""


def find_timestamp(row: Dict[str, Any]) -> Optional[datetime]:
    value = row.get("Timestamp")
    return value if isinstance(value, datetime) else None


def load_form_students(path: Path) -> Dict[str, Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    students: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"stages": {}, "timestamps": [], "stream": None, "name": None})
    for sheet_name, stage in FORM_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        for row in workbook_rows(wb[sheet_name]):
            email = find_email(row)
            if not email:
                continue
            students[email]["stages"][stage] = row
            ts = find_timestamp(row)
            if ts:
                students[email]["timestamps"].append(ts)
            students[email]["stream"] = students[email]["stream"] or as_text(row.get("Stream")) or None
            students[email]["name"] = students[email]["name"] or as_text(row.get("Full Name") or row.get("Name")) or None
    return students


def convert_student(
    email: str,
    student: Dict[str, Any],
    static_lookup: Dict[str, List[StaticQuestion]],
    static_by_id: Dict[str, StaticQuestion],
    ai_lookup: Dict[str, List[AiQuestion]],
    static_aliases: Dict[str, str],
    ai_aliases: Dict[str, str],
) -> Tuple[Optional[Dict[str, Any]], List[Dict[str, str]]]:
    errors: List[Dict[str, str]] = []
    all_responses: Dict[str, Any] = {}
    adaptive_items: List[Dict[str, Any]] = []

    def err(stage: str, question: str, message: str):
        errors.append({"email": email, "stage": stage, "question": question, "message": message})

    # Static non-adaptive sections.
    for stage in ("bigfive", "riasec", "values"):
        row = student["stages"].get(stage)
        if not row:
            continue
        seen: Dict[str, Any] = {}
        for header, answer in row.items():
            if header.strip().lower() in IDENTITY_HEADERS or answer in (None, ""):
                continue
            q = resolve_question(header, static_lookup, static_aliases, static_by_id)
            if not q:
                err(stage, header, "question not mapped")
                continue
            key = normalize_answer(answer).lower()
            score = LIKERT_MAPS[stage].get(key)
            if score is None:
                err(stage, header, f"unknown Likert answer: {answer!r}")
                continue
            if q.id in seen and seen[q.id] != score:
                err(stage, header, f"duplicate question has conflicting answers: {seen[q.id]} vs {score}")
                continue
            seen[q.id] = score
            all_responses[q.id] = score

    # Employability: normal Likert + [Best]/[Worst] pairs.
    row = student["stages"].get("employability")
    if row:
        sjt: Dict[str, Dict[str, str]] = defaultdict(dict)
        for header, answer in row.items():
            if header.strip().lower() in IDENTITY_HEADERS or answer in (None, ""):
                continue
            base, kind = strip_sjt_suffix(header)
            # Google Forms can export helper grid columns such as
            # "Question [specific option]". They are not separate DB questions.
            if "[" in header and "]" in header and kind is None:
                continue
            q = resolve_question(base if kind else header, static_lookup, static_aliases, static_by_id)
            if not q:
                err("employability", header, "question not mapped")
                continue
            if kind:
                val = normalize_answer(answer)
                if kind in sjt[q.id] and sjt[q.id][kind] != val:
                    err("employability", header, f"duplicate SJT {kind} has conflicting answers")
                    continue
                sjt[q.id][kind] = val
            else:
                key = normalize_answer(answer).lower()
                score = LIKERT_MAPS["employability"].get(key)
                if score is None:
                    # If it is a non-Likert SJT-like value, preserve text under
                    # the verified question UUID rather than lose the answer.
                    all_responses[q.id] = normalize_answer(answer)
                else:
                    all_responses[q.id] = score
        for qid, pair in sjt.items():
            if set(pair) != {"best", "worst"}:
                err("employability", qid, f"incomplete SJT pair: {sorted(pair)}")
            all_responses[qid] = {"best": pair.get("best"), "worst": pair.get("worst")}

    # MBA knowledge + aptitude are AI-question JSON UUIDs.
    for stage in ("mba_knowledge", "mba_aptitude"):
        row = student["stages"].get(stage)
        if not row:
            continue
        for header, answer in row.items():
            if header.strip().lower() in IDENTITY_HEADERS or header.strip().lower() == "specialization" or answer in (None, ""):
                continue
            q = resolve_question(header, ai_lookup, ai_aliases)
            if not q:
                err(stage, header, "AI question not mapped")
                continue
            all_responses[q.id] = ai_answer_for_storage(answer, q.options)

    # General Aptitude -> adaptive session/responses/results only.
    row = student["stages"].get("adaptive_aptitude")
    if row:
        sequence = 0
        for header, answer in row.items():
            if header.strip().lower() in IDENTITY_HEADERS or answer in (None, ""):
                continue
            q = resolve_question(header, static_lookup, static_aliases, static_by_id)
            if not q:
                err("adaptive_aptitude", header, "adaptive question not mapped")
                continue
            sequence += 1
            selected = selected_option_key(q.options, answer)
            if selected is None:
                err("adaptive_aptitude", header, f"answer does not match an option: {answer!r}")
                continue
            correct = (q.correct_answer or "").strip('"')
            difficulty = infer_difficulty(q)
            adaptive_items.append({
                "question_id": q.id,
                "question_text": q.text,
                "question_options": q.options,
                "selected_answer": selected,
                "correct_answer": correct,
                "is_correct": selected == correct,
                "response_time_ms": 0,  # unavailable in Google Forms
                "difficulty_at_time": difficulty,
                "subtag": infer_subtag(q),
                "phase": phase_for_sequence(sequence),
                "sequence_number": sequence,
                "explanation": q.description,
            })

    if errors:
        return None, errors

    timestamps = student.get("timestamps") or []
    started_at = min(timestamps) if timestamps else now_utc()
    completed_at = max(timestamps) if timestamps else started_at

    attempt_id = str(uuid.uuid4())
    session_id = str(uuid.uuid4())
    result_id = str(uuid.uuid4())

    total_correct = sum(1 for x in adaptive_items if x["is_correct"])
    total_questions = len(adaptive_items)
    accuracy = (total_correct / total_questions * 100) if total_questions else 0
    by_diff, by_subtag = score_summary(adaptive_items)
    level = aptitude_level_from_accuracy(accuracy)
    tier = tier_from_level(level)
    confidence = confidence_from_accuracy(accuracy)

    difficulty_path = [str(x["difficulty_at_time"]) for x in adaptive_items]
    adaptive_all_responses = []
    for x in adaptive_items:
        adaptive_all_responses.append({
            "phase": x["phase"],
            "subtag": x["subtag"],
            "timestamp": completed_at.isoformat(),
            "is_correct": x["is_correct"],
            "question_id": x["question_id"],
            "selected_answer": x["selected_answer"],
            "sequence_number": x["sequence_number"],
            "response_time_ms": x["response_time_ms"],
            "difficulty_at_time": x["difficulty_at_time"],
        })

    return {
        "email": email,
        "name": student.get("name"),
        "program": student.get("stream"),
        "attempt_id": attempt_id,
        "session_id": session_id,
        "result_id": result_id,
        "started_at": started_at,
        "completed_at": completed_at,
        "all_responses": all_responses,
        "adaptive_items": adaptive_items,
        "adaptive_all_responses": adaptive_all_responses,
        "total_questions": total_questions,
        "total_correct": total_correct,
        "accuracy": accuracy,
        "by_diff": by_diff,
        "by_subtag": by_subtag,
        "difficulty_path": difficulty_path,
        "aptitude_level": level,
        "tier": tier,
        "confidence": confidence,
    }, []


def generate_student_sql(data: Dict[str, Any]) -> str:
    email = data["email"]
    attempt_id = data["attempt_id"]
    session_id = data["session_id"]
    result_id = data["result_id"]
    started = data["started_at"]
    completed = data["completed_at"]

    # The sample learner uses stream_id='college' for MBA/PG assessment.
    learner_context = {
        "rawGrade": "PG",
        "degreeLevel": "Postgraduate",
        "programCode": data.get("program") or "",
        "programName": data.get("program") or "",
        "selectedStream": "college",
        "migrationSource": "google_forms",
    }

    lines = [
        f"-- ============================================================",
        f"-- Learner: {email}",
        f"-- Source: Assessment Answers.xlsx (Google Forms migration)",
        f"-- ============================================================",
        "DO $$",
        "DECLARE",
        "    v_learner_id uuid;",
        "BEGIN",
        f"    SELECT id INTO v_learner_id FROM public.learners WHERE lower(email) = lower({sql_literal(email)}) LIMIT 1;",
        "    IF v_learner_id IS NULL THEN",
        f"        RAISE EXCEPTION 'Learner not found for email: %', {sql_literal(email)};",
        "    END IF;",
        "",
        "    INSERT INTO public.adaptive_aptitude_sessions (",
        "        id, learner_id, grade_level, current_phase, tier, current_difficulty, difficulty_path,",
        "        questions_answered, correct_answers, current_question_index, current_phase_questions,",
        "        provisional_band, status, started_at, updated_at, completed_at, learner_course, all_responses",
        "    ) VALUES (",
        f"        {sql_literal(session_id)}::uuid, v_learner_id, 'postgraduate', 'stability_confirmation', {sql_literal(data['tier'])},",
        f"        {data['adaptive_items'][-1]['difficulty_at_time'] if data['adaptive_items'] else 1}, {sql_smallint_array(data['difficulty_path'])},",
        f"        {data['total_questions']}, {data['total_correct']}, 6, '[]'::jsonb,",
        f"        {data['aptitude_level']}, 'completed', {sql_literal(started)}, {sql_literal(completed)}, {sql_literal(completed)}, NULL,",
        f"        {sql_literal(data['adaptive_all_responses'])}",
        "    );",
        "",
        "    INSERT INTO public.personal_assessment_attempts (",
        "        id, learner_id, stream_id, started_at, completed_at, status, current_section_index,",
        "        current_question_index, section_timings, created_at, updated_at, timer_remaining, elapsed_time,",
        "        grade_level, adaptive_aptitude_session_id, all_responses, aptitude_scores, knowledge_scores,",
        "        aptitude_question_timer, learner_context",
        "    ) VALUES (",
        f"        {sql_literal(attempt_id)}::uuid, v_learner_id, 'college', {sql_literal(started)}, {sql_literal(completed)},",
        "        'completed', 6, 19, '{}'::jsonb,",
        f"        {sql_literal(started)}, {sql_literal(completed)}, NULL, 0, 'college', {sql_literal(session_id)}::uuid,",
        f"        {sql_literal(data['all_responses'])}, NULL, NULL, NULL,",
        f"        {sql_literal(learner_context)}",
        "    );",
        "",
    ]

    for item in data["adaptive_items"]:
        response_id = str(uuid.uuid4())
        lines += [
            "    INSERT INTO public.adaptive_aptitude_responses (",
            "        id, session_id, question_id, selected_answer, is_correct, response_time_ms, difficulty_at_time,",
            "        subtag, phase, sequence_number, answered_at, question_text, question_options, correct_answer, explanation",
            "    ) VALUES (",
            f"        {sql_literal(response_id)}::uuid, {sql_literal(session_id)}::uuid, {sql_literal(item['question_id'])}::uuid,",
            f"        {sql_literal(item['selected_answer'])}, {sql_literal(item['is_correct'])}, {item['response_time_ms']},",
            f"        {item['difficulty_at_time']}, {sql_literal(item['subtag'])}, {sql_literal(item['phase'])}, {item['sequence_number']},",
            f"        {sql_literal(completed)}, {sql_literal(item['question_text'])}, {sql_literal(item['question_options'])},",
            f"        {sql_literal(item['correct_answer'])}, {sql_literal(item['explanation'])}",
            "    );",
        ]

    metadata = {
        "duplicateValidation": {"isValid": True, "duplicates": []},
        "migrationSource": "google_forms",
        "responseTimeUnavailable": True,
    }
    lines += [
        "",
        "    INSERT INTO public.adaptive_aptitude_results (",
        "        id, session_id, learner_id, aptitude_level, confidence_tag, tier, total_questions, total_correct,",
        "        overall_accuracy, accuracy_by_difficulty, accuracy_by_subtag, difficulty_path, path_classification,",
        "        average_response_time_ms, grade_level, completed_at, created_at, metadata",
        "    ) VALUES (",
        f"        {sql_literal(result_id)}::uuid, {sql_literal(session_id)}::uuid, v_learner_id, {data['aptitude_level']},",
        f"        {sql_literal(data['confidence'])}, {sql_literal(data['tier'])}, {data['total_questions']}, {data['total_correct']},",
        f"        {data['accuracy']}, {sql_literal(data['by_diff'])}, {sql_literal(data['by_subtag'])}, {sql_smallint_array(data['difficulty_path'])},",
        f"        'stable', 0, 'postgraduate', {sql_literal(completed)}, {sql_literal(completed)}, {sql_literal(metadata)}",
        "    );",
        "END $$;",
        "",
    ]
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("answers", nargs="?", default="Assessment Answers.xlsx")
    parser.add_argument("--question-bank", default="question_bank.xlsx")
    parser.add_argument("--output-dir", default="output")
    parser.add_argument("--ai-sql", default="career_assessment_ai_questions_rows.sql", help="Optional SQL dump containing MBA aptitude question JSON")
    parser.add_argument("--allow-partial", action="store_true", help="Generate SQL for valid students even when other students fail validation")
    args = parser.parse_args()

    answers = Path(args.answers)
    bank = Path(args.question_bank)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if not answers.exists():
        print(f"ERROR: answers file not found: {answers}", file=sys.stderr)
        return 2
    if not bank.exists():
        print(f"ERROR: question bank not found: {bank}", file=sys.stderr)
        return 2

    static_lookup, static_by_id, ai_lookup = load_reference_bank(bank)
    ai_sql = Path(args.ai_sql) if args.ai_sql else None
    load_ai_questions_from_sql(ai_sql, ai_lookup)
    static_aliases = build_alias_index(static_lookup.keys())
    ai_aliases = build_alias_index(ai_lookup.keys())
    students = load_form_students(answers)

    valid: List[Dict[str, Any]] = []
    all_errors: List[Dict[str, str]] = []
    for email, student in sorted(students.items()):
        data, errors = convert_student(email, student, static_lookup, static_by_id, ai_lookup, static_aliases, ai_aliases)
        if errors:
            all_errors.extend(errors)
        elif data:
            valid.append(data)

    report_path = out_dir / "validation_report.csv"
    with report_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["email", "stage", "question", "message"])
        writer.writeheader()
        writer.writerows(all_errors)

    if all_errors and not args.allow_partial:
        print(f"Validation failed: {len(all_errors)} issue(s).")
        print(f"Review: {report_path}")
        print("No SQL generated. Re-run with --allow-partial only if you intentionally want valid students only.")
        return 1

    print(f"Students found: {len(students)}")
    print(f"Valid students: {len(valid)}")
    print(f"Validation issues: {len(all_errors)}")
    print(f"Report: {report_path}")

    while True:
        raw = input(
            f"How many valid students do you want to generate seed files for? "
            f"Enter 0 for all {len(valid)}: "
        ).strip()
        try:
            count = int(raw)
        except ValueError:
            print("Enter a whole number, for example 0, 1, 5, 10.")
            continue
        if count < 0:
            print("Count cannot be negative.")
            continue
        if count > len(valid):
            print(f"Only {len(valid)} valid students are available.")
            continue
        break

    selected = valid if count == 0 else valid[:count]
    seeds_dir = out_dir / "seeds"
    seeds_dir.mkdir(parents=True, exist_ok=True)

    for data in selected:
        local_part = data["email"].split("@", 1)[0]
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", local_part).strip("._") or "learner"
        seed_path = seeds_dir / f"{safe_name}_assessment_seed.sql"
        with seed_path.open("w", encoding="utf-8") as f:
            f.write("-- Generated Google Forms assessment migration\n")
            f.write(f"-- Learner: {data['email']}\n")
            f.write("-- Target tables: personal_assessment_attempts, adaptive_aptitude_sessions, adaptive_aptitude_responses, adaptive_aptitude_results\n")
            f.write("-- Existing learner and question UUIDs are reused.\n")
            f.write("-- response_time_ms/average_response_time_ms are 0 because Google Forms did not record per-question timing.\n\n")
            f.write("BEGIN;\n\n")
            f.write(generate_student_sql(data))
            f.write("\nCOMMIT;\n")

    print(f"Seed files generated: {len(selected)}")
    print(f"Seeds directory: {seeds_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
